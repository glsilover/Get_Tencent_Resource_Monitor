import csv
import logging
import openpyxl
import os

logger = logging.getLogger(__name__)


def safe_value(value):
    # 如果值不是整数或浮点数，则将其转换为字符串
    if not isinstance(value, (int, float)):
        return str(value)
    return value


def list_to_excel(list_data, file_path, sheet_name):
    try:
        # 检查Excel文件是否存在
        if os.path.exists(file_path):
            workbook = openpyxl.load_workbook(file_path)
            if sheet_name in workbook.sheetnames:
                # 如果sheet存在，则覆盖
                sheet = workbook[sheet_name]
                workbook.remove(sheet)
                workbook.create_sheet(sheet_name, 0)
                sheet = workbook[sheet_name]
            else:
                # 如果sheet不存在，则创建
                sheet = workbook.create_sheet(sheet_name)
        else:
            # 如果文件不存在，则创建新的工作簿
            workbook = openpyxl.Workbook()
            sheet = workbook.active
            sheet.title = sheet_name

        # 找出拥有最多键的字典
        max_keys_dict = max(list_data, key=lambda x: len(x.keys()))
        columns = list(max_keys_dict.keys())

        # 写入列名和数据
        sheet.append(columns)
        for item in list_data:
            row = [safe_value(item.get(col, None)) for col in columns]
            sheet.append(row)

        # 保存Excel文件
        workbook.save(file_path)

        logger.debug(f'数据已写入Excel文件【{file_path}】')

    except Exception as e:
        logger.error(f'写入Excel文件【{file_path}】失败', exc_info=True)


def excel_to_list(file_path, sheet_name):
    # 检查文件是否存在
    if not os.path.exists(file_path):
        raise FileNotFoundError(f'文件【{file_path}】未找到')

    # 加载Excel文件和指定的工作表
    workbook = openpyxl.load_workbook(file_path)
    if sheet_name not in workbook.sheetnames:
        raise ValueError(f'工作表【{sheet_name}】在文件【{file_path}】中不存在')
    sheet = workbook[sheet_name]

    # 读取列名
    columns = []
    for cell in sheet[1]:  # 假设第一行是列名
        columns.append(cell.value)

    # 读取数据行
    list_data = []
    for row in sheet.iter_rows(min_row=2, values_only=True):  # 跳过第一行的列名
        row_data = dict(zip(columns, row))
        list_data.append(row_data)

    return list_data
